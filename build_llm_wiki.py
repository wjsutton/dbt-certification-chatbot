#!/usr/bin/env python3
"""Incremental scaffolder for the dbt cert LLM Wiki.

Idempotent: re-running only rewrites a page when its underlying source data
changes. Concept pages keep agent-written prose; the builder edits only the
marked auto-blocks. log.md is append-only. A .build-state.json tracks raw-file
hashes so changed sources (whose summaries may be stale) are reported.

Usage:
    pip install openpyxl
    python build_llm_wiki.py
    python build_llm_wiki.py --raw raw --xlsx dbt_cert_content_mapping.xlsx --out wiki
"""
from __future__ import annotations
import argparse, csv, datetime as dt, hashlib, json, re
from pathlib import Path
from urllib.parse import urlparse

TODAY = dt.date.today().isoformat()
A_SUB_S, A_SUB_E = "<!-- dbtwiki:auto:subtopics -->", "<!-- /dbtwiki:auto:subtopics -->"
A_SRC_S, A_SRC_E = "<!-- dbtwiki:auto:sources -->", "<!-- /dbtwiki:auto:sources -->"
STUB_SENTINEL = "_TODO (agent): 2–4 paragraph synthesis"

def slug_for(url):
    p = urlparse(url); host = "getdbt" if p.netloc == "www.getdbt.com" else "docs"
    path = re.sub(r"[^A-Za-z0-9_.-]", "-", p.path.strip("/").replace("/", "__"))
    return f"{host}__{path or 'index'}.md"

def summary_seg(url):
    parts = urlparse(url).path.rstrip("/").split("/")
    last = parts[-1] or "index"
    if last[:1].isdigit() and len(parts) >= 2:   # avoid collisions like .../1-guide-overview
        return parts[-2] + "-" + last
    return last

def read_mapping(xlsx):
    from openpyxl import load_workbook
    ws = load_workbook(xlsx, read_only=True, data_only=True)["Topic Mapping"]
    rows, domain, sub = [], "", ""
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or all(c is None for c in r[:5]): continue
        d, s, typ, title, url = (list(r) + [None]*5)[:5]
        if d: domain = str(d).strip()
        if s: sub = str(s).strip()
        if not url: continue
        rows.append({"domain": domain, "subtopic": sub,
                     "type": str(typ).strip() if typ else "",
                     "title": str(title).strip() if title else "", "url": str(url).strip()})
    return rows

def read_manifest(raw):
    man = raw / "_manifest.tsv"; out = {}
    if man.exists():
        for row in csv.DictReader(man.open(encoding="utf-8"), delimiter="\t"):
            out[row["source_url"]] = row
    return out

def dslugf(d):
    return re.sub(r"[^a-z0-9]+", "-", d.lower().replace("—","-").replace("&","and")).strip("-")
def title_for(r, m):
    mt = (m.get(r["url"]) or {}).get("title") or ""
    if mt and "__" not in mt and not mt.endswith(".md"):  # clean manifest title
        return mt
    return r["title"] or mt or slug_for(r["url"])         # fall back to mapping title
def uniq(rows):
    seen, out = set(), []
    for r in rows:
        if r["url"] not in seen: seen.add(r["url"]); out.append(r)
    return out
def nest(rows):
    n = {}
    for r in rows:
        dom = n.setdefault(r["domain"], []); e = next((e for e in dom if e[0]==r["subtopic"]), None)
        if e is None: e = (r["subtopic"], []); dom.append(e)
        if r["url"] not in {s["url"] for s in e[1]}: e[1].append(r)
    return n

def link(r, m, raw_rel, wiki):
    """Link to the source summary if it exists, else to the raw file."""
    seg = summary_seg(r["url"])
    if (wiki / "sources" / f"{seg}.md").exists():
        return f"[{title_for(r,m)}](../sources/{seg}.md)"
    return f"[{title_for(r,m)}]({raw_rel}/{slug_for(r['url'])})"

def gen_subtopics(subs, m, raw_rel, wiki):
    out = ["## Sub-topics assessed", ""]
    for st, srcs in subs:
        out.append(f"- **{st}** — {' · '.join(link(s,m,raw_rel,wiki) for s in srcs)}")
    return "\n".join(out)

def gen_sources(sources, m, raw_rel, wiki, root, raw_name):
    out = ["## Source material", ""]
    for r in sources:
        sg = slug_for(r["url"]); seg = summary_seg(r["url"])
        present = (root / raw_name / sg).exists()
        summ = f" · summary: [{seg}](../sources/{seg}.md)" if (wiki/"sources"/f"{seg}.md").exists() else ""
        mark = "" if present else "  _(run extractor to create)_"
        out.append(f"- [{title_for(r,m)}]({raw_rel}/{sg}){summ} · [original]({r['url']}) · `{r['type']}`{mark}")
    return "\n".join(out)

def fm(title, **e):
    L = ["---", 'title: "' + title.replace('"', "'") + '"']
    for k, v in e.items():
        L.append(f"{k}: [{', '.join(v)}]" if isinstance(v, list) else f"{k}: {v}")
    return "\n".join(L) + "\n---\n"

def set_updated(text, date):
    if re.search(r"^updated:.*$", text, flags=re.M):
        return re.sub(r"^updated:.*$", f"updated: {date}", text, count=1, flags=re.M)
    return text

def replace_between(text, start, end, new_inner):
    pat = re.compile(re.escape(start) + r".*?" + re.escape(end), re.DOTALL)
    return pat.sub(start + "\n" + new_inner + "\n" + end, text, count=1)

def concept_template(domain, sub_block, src_block):
    return (fm(domain, tags=["exam-domain","concept"], status="stub", updated=TODAY)
        + f"\n# {domain}\n\n"
        + f"{A_SUB_S}\n{sub_block}\n{A_SUB_E}\n\n"
        + "## Synthesis\n\n_TODO (agent): 2–4 paragraph synthesis of how these "
          "sub-topics fit together, written from the sources below._\n\n"
        + "## Related pages\n\n_TODO (agent): link to overlapping concept pages "
          "(e.g. state ↔ defer ↔ clone)._\n\n"
        + f"{A_SRC_S}\n{src_block}\n{A_SRC_E}\n")

def write_if_changed(path, content, report, label):
    if path.exists() and path.read_text(encoding="utf-8") == content:
        report["unchanged"].append(label); return False
    existed = path.exists()
    path.write_text(content, encoding="utf-8")
    report["updated" if existed else "created"].append(label); return True


SCHEMA = """# CLAUDE.md — dbt Cert Wiki operating manual

You maintain an **LLM Wiki** (Karpathy pattern) for the **dbt Analytics
Engineering Certification (dbt Core 1.11)**. This is your schema and playbook.

## Layers
- `raw/` — immutable sources (one md per page, from extract_dbt_docs.py). Read only.
- `wiki/` — your knowledge base: concepts/, sources/, notes/, index.md, log.md, overview.md.
- `CLAUDE.md` — these rules. Co-evolve deliberately.

## Incremental builds (important)
`build_llm_wiki.py` is **incremental**. It only rewrites a page when the
underlying source data changes, and it never overwrites prose you write.
- Concept pages have auto-blocks delimited by `<!-- dbtwiki:auto:subtopics -->`
  and `<!-- dbtwiki:auto:sources -->`. The builder edits **only** inside those
  markers. Write your synthesis between the markers (Synthesis / Related pages)
  and it is preserved across rebuilds.
- `sources/` and `notes/` are never touched by the builder.
- `log.md` is append-only. `index.md` / `overview.md` are regenerated only when
  their content changes.
- A raw file whose content changed is reported as a **stale summary** to re-ingest.

## Operations
- **Ingest**: read a `raw/` file; write `sources/<seg>.md` (summary, sub-topics,
  key tokens, gotchas); fill the **Synthesis** of the relevant `concepts/<domain>.md`
  (between the markers) and bump its `status` (stub -> draft -> done). Re-run the
  builder to refresh links; append happens automatically.
- **Query (chatbot)**: read index.md, drill into concept + source pages, answer
  ONLY from wiki/raw, always cite the source page + original dbt URL, use exact
  tokens (e.g. `--empty`, `--sample`, `state:modified`, `grants`). File reusable
  answers as `notes/<topic>.md`.
- **Lint**: flag contradictions, version drift (target dbt Core 1.11), orphans,
  uncovered sub-topics, empty stubs, and stale summaries.

## Scope
`dbt_cert_content_mapping.xlsx` is the canonical topic->source map. Add sources
there and re-run the builder.
"""

def ensure_schema(root):
    fp = root / "CLAUDE.md"
    if not fp.exists():
        fp.write_text(SCHEMA, encoding="utf-8"); return "created CLAUDE.md"
    return None

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--raw", default="raw"); ap.add_argument("--out", default="wiki")
    ap.add_argument("--xlsx", default="dbt_cert_content_mapping.xlsx")
    a = ap.parse_args()
    xlsx, raw, wiki = Path(a.xlsx), Path(a.raw), Path(a.out)
    if not xlsx.exists(): raise SystemExit(f"missing {xlsx}")
    wiki.mkdir(parents=True, exist_ok=True)
    (wiki/"concepts").mkdir(exist_ok=True); (wiki/"sources").mkdir(exist_ok=True); (wiki/"notes").mkdir(exist_ok=True)
    root = wiki.parent if str(wiki.parent) else Path("."); raw_rel = "../" + raw.name
    _sc = ensure_schema(root)
    rows, man = read_mapping(xlsx), read_manifest(raw)
    rep = {"created": [], "updated": [], "unchanged": [], "skipped": []}
    if _sc: rep["created"].append("CLAUDE.md")

    # ---- concept pages (incremental, marker-aware) ----
    summary = []
    for domain, subs in nest(rows).items():
        ds = dslugf(domain); fp = wiki/"concepts"/f"{ds}.md"
        sources = uniq([r for r in rows if r["domain"] == domain])
        sub_block = gen_subtopics(subs, man, raw_rel, wiki)
        src_block = gen_sources(sources, man, raw_rel, wiki, root, raw.name)
        summary.append((ds, domain, len(sources), len([s[0] for s in subs if s[0]])))
        if not fp.exists():
            write_if_changed(fp, concept_template(domain, sub_block, src_block), rep, f"concepts/{ds}.md")
            continue
        cur = fp.read_text(encoding="utf-8")
        if A_SUB_S in cur and A_SRC_S in cur:                      # managed page → refresh auto blocks only
            new = replace_between(cur, A_SUB_S, A_SUB_E, sub_block)
            new = replace_between(new, A_SRC_S, A_SRC_E, src_block)
            if new != cur:
                new = set_updated(new, TODAY)
                write_if_changed(fp, new, rep, f"concepts/{ds}.md")
            else:
                rep["unchanged"].append(f"concepts/{ds}.md")
        elif STUB_SENTINEL in cur and re.search(r"^status:\s*stub", cur, re.M):  # untouched stub → upgrade
            write_if_changed(fp, concept_template(domain, sub_block, src_block), rep, f"concepts/{ds}.md (upgraded)")
        else:                                                     # hand-written, no markers → never clobber
            rep["skipped"].append(f"concepts/{ds}.md (no markers; left untouched)")

    # ---- overview (catalog, no human content) ----
    ov = [fm("Overview — dbt Analytics Engineering Certification wiki", tags=["overview"], updated=TODAY),
          "\n# Overview\n", "Compounding knowledge base for the **dbt Analytics Engineering "
          "Certification (dbt Core 1.11)**, Karpathy LLM-Wiki pattern. Raw docs in `../raw/` "
          "(immutable); this `wiki/` is agent-maintained per `../CLAUDE.md`.\n", "## Exam structure\n"]
    for ds, d, ns, nb in summary: ov.append(f"- [{d}](concepts/{ds}.md) — {nb} sub-topics, {ns} sources")
    ov.append("\n## How to use\n- Browse from [index.md](index.md).\n- Ask the agent; it reads the "
              "index, drills into concept + source pages, answers with citations.\n- Good answers get filed back.\n")
    write_if_changed(wiki/"overview.md", "\n".join(ov), rep, "overview.md")

    # ---- index (catalog; Source summaries auto-listed from sources/) ----
    ix = [fm("Index", tags=["index"], updated=TODAY), "\n# Index\n",
          "Catalog of the wiki. Read this first when answering a query.\n",
          "## Start here\n- [Overview](overview.md)\n- [Change log](log.md)\n",
          "## Contents (by exam domain, sub-topic & source)\n"]
    dbd = {d: ds for ds, d, _, _ in summary}
    for domain, subs in nest(rows).items():
        ix.append(f"### [{domain}](concepts/{dbd[domain]}.md)\n")
        for st, srcs in subs: ix.append(f"- **{st}** — {' · '.join(link(s,man,raw_rel,wiki) for s in srcs)}")
        ix.append("")
    summaries = sorted((wiki/"sources").glob("*.md"))
    if summaries:
        ix.append("## Source summaries (ingested)\n")
        for sp in summaries:
            t = sp.read_text(encoding="utf-8")
            title = (re.search(r'title:\s*"(.*?)"', t) or [None, sp.stem])[1] if re.search(r'title:', t) else sp.stem
            st = (re.search(r'status:\s*(\w+)', t) or [None, "?"])[1] if re.search(r'status:', t) else "?"
            ix.append(f"- [{title}](sources/{sp.name}) — `{st}`")
        ix.append("")
    src_all = uniq(rows)
    ix.append(f"## Raw sources ({len(src_all)})\n")
    for r in sorted(src_all, key=lambda x: x["url"]):
        ix.append(f"- [{title_for(r,man)}]({raw_rel}/{slug_for(r['url'])}) · `{r['type']}` · [source]({r['url']})")
    ix.append("")
    write_if_changed(wiki/"index.md", "\n".join(ix), rep, "index.md")

    # ---- raw-hash state: detect changed sources whose summaries may be stale ----
    state_fp = wiki/".build-state.json"
    state = json.loads(state_fp.read_text()) if state_fp.exists() else {"raw_hashes": {}}
    old = state.get("raw_hashes", {}); new = {}
    for f in sorted(raw.glob("*.md")):
        new[f.name] = hashlib.sha1(f.read_bytes()).hexdigest()
    changed = [n for n in new if n in old and new[n] != old[n]]
    added = [n for n in new if n not in old]
    stale = []
    for n in changed:
        seg = n.split("__")[-1].replace(".md", "")
        if (wiki/"sources"/f"{seg}.md").exists(): stale.append(f"sources/{seg}.md (raw changed)")
    state["raw_hashes"] = new; state["last_run"] = TODAY
    state_fp.write_text(json.dumps(state, indent=2), encoding="utf-8")

    # ---- log: append-only, only when something changed ----
    touched = rep["created"] + rep["updated"]
    log = wiki/"log.md"
    if not log.exists():
        log.write_text(fm("Log", tags=["log"]) + "\n# Log\n\nAppend-only timeline.\n", encoding="utf-8")
    if touched or added or changed:
        bits = []
        if rep["created"]: bits.append(f"created {len(rep['created'])}")
        if rep["updated"]: bits.append(f"updated {len(rep['updated'])}")
        if added: bits.append(f"{len(added)} new raw")
        if changed: bits.append(f"{len(changed)} changed raw")
        entry = f"\n## [{TODAY}] build | {', '.join(bits)}\n"
        for x in rep["created"]: entry += f"- created {x}\n"
        for x in rep["updated"]: entry += f"- updated {x}\n"
        for x in stale: entry += f"- (!) stale: {x} — re-ingest\n"
        with log.open("a", encoding="utf-8") as fh: fh.write(entry)

    # ---- report ----
    print("Incremental build:")
    for k in ("created", "updated", "unchanged", "skipped"):
        if rep[k]: print(f"  {k}: {len(rep[k])}")
        for x in rep[k]:
            if k in ("created", "updated", "skipped"): print(f"      - {x}")
    if added: print(f"  new raw sources: {len(added)}")
    if changed: print(f"  changed raw sources: {changed}")
    if stale:
        print("  STALE summaries (re-ingest):")
        for s in stale: print(f"      - {s}")
    if not touched and not added and not changed:
        print("  nothing changed — no files written.")


if __name__ == "__main__":
    main()
