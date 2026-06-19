#!/usr/bin/env python3
"""Question-bank coverage vs the certification blueprint.
Shows item counts per domain & format, and lists exam sub-topics that have
NO items yet — i.e. what to develop next.

Usage:  python coverage.py
"""
import json, collections
from pathlib import Path
HERE=Path(__file__).resolve().parent; PROJ=HERE.parent
def main():
    items=json.loads((HERE/"questions.json").read_text())["items"]
    bp=json.loads((HERE/"blueprints"/"exam-65.json").read_text())
    by_dom=collections.Counter(i["domain"] for i in items)
    by_fmt=collections.Counter(i["format"] for i in items)
    have_sub=collections.Counter((i["domain"], i["subtopic"]) for i in items)
    print("=== Items by domain (have / blueprint target) ===")
    for d,tgt in bp["domain_counts"].items():
        print(f"  {by_dom.get(d,0):>3} / {tgt:<3}  {d}")
    print("\n=== Items by format ===")
    for f,c in sorted(by_fmt.items()): print(f"  {c:>3}  {f}")
    # sub-topic gaps from the mapping spreadsheet (optional)
    try:
        from openpyxl import load_workbook
        ws=load_workbook(PROJ/"dbt_cert_content_mapping.xlsx",read_only=True,data_only=True)["Topic Mapping"]
        subs, dom, sub = [], "", ""
        for r in ws.iter_rows(min_row=3, values_only=True):
            if not r or all(c is None for c in r[:5]): continue
            d,s=(list(r)+[None]*5)[:2]
            if d: dom=str(d).strip()
            if s: sub=str(s).strip()
            if sub and (dom,sub) not in subs: subs.append((dom,sub))
        gaps=[(d,s) for (d,s) in subs if have_sub.get((d,s),0)==0]
        print(f"\n=== Sub-topics with NO items yet ({len(gaps)} of {len(subs)}) ===")
        cur=None
        for d,s in gaps:
            if d!=cur: print(f"  {d}"); cur=d
            print(f"      - {s}")
    except Exception as e:
        print("\n(install openpyxl to see sub-topic gaps:", e, ")")
if __name__=="__main__": main()
